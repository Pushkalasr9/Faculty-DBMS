import openpyxl
import pandas as pd



theFile = openpyxl.load_workbook('C:/Users/Priyanshu Kumar/Desktop/My Codes/Python\django\SEPM\FDBMS\data\FacultyDetails.xlsx')
currentSheet = theFile['Sheet1']

cols = ['B', 'C', 'D', 'E', 'F']

def getdetails(key):
    data={}
    details = {}
    users = []
    for i in range(2, currentSheet.max_row+1):
        if key == currentSheet[f'F{i}'].value:
            details["id"]= str(currentSheet[f'B{i}'].value)
            details["name"]= str(currentSheet[f'C{i}'].value)
            details["pos"]= str(currentSheet[f'D{i}'].value)
            details["phone"]= str(currentSheet[f'E{i}'].value)
            details["email"]= str(currentSheet[f'F{i}'].value)
            users.append(details)
    data['users']=users
    return data
    
def alldetails():
    data = {}
    c = 0
    users=[]
    ids=[]
    for i in range(2, currentSheet.max_row+1):
        details = {}
        details["id"]= str(currentSheet[f'B{i}'].value)
        details["name"]= str(currentSheet[f'C{i}'].value)
        details["pos"]= str(currentSheet[f'D{i}'].value)
        details["phone"]= str(currentSheet[f'E{i}'].value)
        details["email"]= str(currentSheet[f'F{i}'].value)
        ids.append(str(f"{currentSheet[f'B{i}'].value}"))
        users.append(details)
        c+=1
    data['users']=users
    data['ids']=ids
    data['max']=currentSheet.max_row
    return data

def store(data):
    sno = currentSheet.max_row
    id = data['Faculty ID']
    name = data['Faculty Name']
    pos = data['Position']
    phone = data['Contact Number']
    email = data['Email ID']
    #print(data)
    data['S.No'] = pd.Series([sno])
    data['Faculty ID'] = pd.Series([id])
    data['Faculty Name'] = pd.Series([name])
    data['Position'] = pd.Series([pos])
    data['Contact Number'] = pd.Series([phone])
    data['Email ID'] = pd.Series([email])
    #print(data)
    df = pd.read_excel('C:/Users/Priyanshu Kumar/Desktop/My Codes/Python\django\SEPM\FDBMS\data\FacultyDetails.xlsx', sheet_name='Sheet1', engine='openpyxl')
    new_data = pd.DataFrame(data)
    df = pd.concat([df, new_data], ignore_index=True)

    # Write the updated DataFrame to the Excel file
    with pd.ExcelWriter('C:/Users/Priyanshu Kumar/Desktop/My Codes/Python\django\SEPM\FDBMS\data\FacultyDetails.xlsx', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
    
    
'''data = {
    'Faculty ID': pd.Series(['123']),
    'Faculty Name': pd.Series(['Priyanshu']),
    'Position': pd.Series(['Student']),
    'Contact Number': pd.Series(['1234567890']),
    'Email ID': pd.Series(['r'])
}

print(data)'''